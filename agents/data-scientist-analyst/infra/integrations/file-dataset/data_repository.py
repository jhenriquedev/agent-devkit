#!/usr/bin/env python3
"""Local dataset repository for the Data Scientist Analyst agent."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any


class DataScientistError(RuntimeError):
    """Raised for user-facing dataset analysis errors."""


@dataclass(frozen=True)
class Dataset:
    source: str
    name: str
    format: str
    rows: list[dict[str, str]]
    columns: list[str]
    sha256: str


CPF_RE = re.compile(r"\b\d{3}\.?\d{3}\.?\d{3}-?\d{2}\b")
CNPJ_RE = re.compile(r"\b\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}\b")
EMAIL_RE = re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.IGNORECASE)
PHONE_RE = re.compile(r"(?:\+?55\s*)?(?:\(?\d{2}\)?\s*)?\d{4,5}[-\s]?\d{4}")

SENSITIVE_NAME_PATTERNS = {
    "cpf": "cpf",
    "cnpj": "cnpj",
    "email": "email",
    "e-mail": "email",
    "telefone": "phone",
    "phone": "phone",
    "celular": "phone",
    "nome": "person_name",
    "name": "person_name",
}


class DataRepository:
    """Read-only local tabular analysis repository."""

    def ingest_dataset(self, source: str) -> dict[str, Any]:
        dataset = self.load_dataset(source)
        return {
            "dataset": self.dataset_summary(dataset),
            "columns": dataset.columns,
            "sample_rows": [mask_row(row) for row in dataset.rows[:5]],
        }

    def inspect_dataset_schema(self, source: str) -> dict[str, Any]:
        dataset = self.load_dataset(source)
        profile = self.profile_dataset(source)
        return {
            "dataset": profile["dataset"],
            "schema": profile["schema"],
            "columns": {
                column: {
                    "inferred_type": details["inferred_type"],
                    "missing_count": details["missing_count"],
                    "unique_count": details["unique_count"],
                    "sample_values": details["sample_values"],
                }
                for column, details in profile["columns"].items()
            },
            "sensitive_data": profile["sensitive_data"],
        }

    def profile_dataset(self, source: str) -> dict[str, Any]:
        dataset = self.load_dataset(source)
        column_profiles = {
            column: profile_column(column, [row.get(column, "") for row in dataset.rows])
            for column in dataset.columns
        }
        duplicate_row_count = count_duplicate_rows(dataset.rows, dataset.columns)
        sensitive = detect_sensitive(dataset)
        probable_keys = infer_probable_keys(dataset, column_profiles)
        return {
            "dataset": self.dataset_summary(dataset),
            "schema": {
                "columns": dataset.columns,
                "probable_keys": probable_keys,
                "types": {
                    column: details["inferred_type"]
                    for column, details in column_profiles.items()
                },
            },
            "quality": {
                "duplicate_row_count": duplicate_row_count,
                "empty_row_count": count_empty_rows(dataset.rows),
                "columns_with_missing_values": [
                    column
                    for column, details in column_profiles.items()
                    if details["missing_count"] > 0
                ],
                "quality_score": quality_score(
                    row_count=len(dataset.rows),
                    column_profiles=column_profiles,
                    duplicate_row_count=duplicate_row_count,
                ),
            },
            "columns": column_profiles,
            "sensitive_data": sensitive,
        }

    def detect_sensitive_data(self, source: str) -> dict[str, Any]:
        dataset = self.load_dataset(source)
        return {
            "dataset": self.dataset_summary(dataset),
            "sensitive_data": detect_sensitive(dataset),
        }

    def run_exploratory_analysis(
        self,
        source: str,
        target_column: str | None = None,
        segment_column: str | None = None,
    ) -> dict[str, Any]:
        profile = self.profile_dataset(source)
        outliers = self.detect_outliers(source=source, columns=None, method="iqr", threshold=3.0)
        correlations = self.analyze_correlation(
            source=source,
            columns=None,
            target_column=target_column,
        )
        segments = (
            self.segment_data(
                source=source,
                segment_column=segment_column,
                metric_column=target_column,
                max_segments=10,
            )
            if segment_column
            else {"summary": {"segment_count": 0}, "segments": []}
        )
        hypotheses = build_hypotheses(profile, outliers, correlations, segments, target_column)
        return {
            "dataset": profile["dataset"],
            "schema": profile["schema"],
            "quality": profile["quality"],
            "numeric_columns": [
                column
                for column, details in profile["columns"].items()
                if details["inferred_type"] == "number"
            ],
            "categorical_columns": [
                column
                for column, details in profile["columns"].items()
                if details["inferred_type"] == "text"
            ],
            "outliers": outliers,
            "correlations": correlations,
            "segments": segments,
            "hypotheses": hypotheses,
            "limitations": [
                "Analise exploratoria nao demonstra causalidade.",
                "Resultados dependem da qualidade e cobertura da fonte informada.",
            ],
        }

    def detect_outliers(
        self,
        source: str,
        columns: list[str] | None = None,
        method: str = "iqr",
        threshold: float = 3.0,
    ) -> dict[str, Any]:
        dataset = self.load_dataset(source)
        numeric_columns = resolve_numeric_columns(dataset, columns)
        result = {}
        for column in numeric_columns:
            values = numeric_values_with_rows(dataset.rows, column)
            if len(values) < 4:
                result[column] = {
                    "method": method,
                    "outlier_count": 0,
                    "outlier_ratio": 0.0,
                    "outliers": [],
                    "limits": {},
                }
                continue
            result[column] = detect_column_outliers(values, method, threshold)
        return {
            "dataset": self.dataset_summary(dataset),
            "method": method,
            "threshold": threshold,
            "columns": result,
        }

    def analyze_correlation(
        self,
        source: str,
        columns: list[str] | None = None,
        target_column: str | None = None,
    ) -> dict[str, Any]:
        dataset = self.load_dataset(source)
        numeric_columns = resolve_numeric_columns(dataset, columns)
        pairs = []
        for left_index, left in enumerate(numeric_columns):
            for right in numeric_columns[left_index + 1 :]:
                paired = paired_numeric_values(dataset.rows, left, right)
                if len(paired) < 2:
                    continue
                coefficient = pearson_correlation([item[0] for item in paired], [item[1] for item in paired])
                if coefficient is None:
                    continue
                pairs.append(
                    {
                        "left": left,
                        "right": right,
                        "correlation": round(coefficient, 6),
                        "strength": correlation_strength(coefficient),
                        "sample_size": len(paired),
                    }
                )
        pairs.sort(key=lambda item: abs(item["correlation"]), reverse=True)
        target_pairs = []
        if target_column:
            for item in pairs:
                if item["left"] == target_column:
                    target_pairs.append({"column": item["right"], **item})
                elif item["right"] == target_column:
                    target_pairs.append({"column": item["left"], **item})
            target_pairs.sort(key=lambda item: abs(item["correlation"]), reverse=True)
        return {
            "dataset": self.dataset_summary(dataset),
            "numeric_columns": numeric_columns,
            "target_column": target_column,
            "top_correlations": pairs[:20],
            "target_correlations": target_pairs[:20],
            "limitations": [
                "Correlacao mede associacao linear, nao causalidade.",
                "Colunas com baixa variancia ou poucos pares validos podem ser omitidas.",
            ],
        }

    def segment_data(
        self,
        source: str,
        segment_column: str,
        metric_column: str | None = None,
        max_segments: int = 20,
    ) -> dict[str, Any]:
        dataset = self.load_dataset(source)
        if segment_column not in dataset.columns:
            raise DataScientistError(f"segment column not found: {segment_column}")
        if metric_column and metric_column not in dataset.columns:
            raise DataScientistError(f"metric column not found: {metric_column}")
        grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
        for row in dataset.rows:
            segment = row.get(segment_column, "") or "(empty)"
            grouped[segment].append(row)
        segments = []
        total = len(dataset.rows)
        for segment, rows in grouped.items():
            item: dict[str, Any] = {
                "segment": mask_value(segment_column, segment),
                "row_count": len(rows),
                "row_ratio": round(len(rows) / total, 6) if total else 0.0,
            }
            if metric_column:
                metric_values = [float(value) for value in (to_decimal(row.get(metric_column, "")) for row in rows) if value is not None]
                item["metric"] = summarize_metric(metric_values)
            segments.append(item)
        segments.sort(key=lambda item: item["row_count"], reverse=True)
        return {
            "dataset": self.dataset_summary(dataset),
            "segment_column": segment_column,
            "metric_column": metric_column,
            "summary": {
                "segment_count": len(segments),
                "returned_segment_count": min(len(segments), max_segments),
                "total_rows": total,
            },
            "segments": segments[:max_segments],
        }

    def analyze_time_series(
        self,
        source: str,
        date_column: str,
        metric_column: str,
        granularity: str = "day",
    ) -> dict[str, Any]:
        dataset = self.load_dataset(source)
        series = aggregate_time_series(dataset, date_column, metric_column, granularity)
        return {
            "dataset": self.dataset_summary(dataset),
            "date_column": date_column,
            "metric_column": metric_column,
            "granularity": granularity,
            "summary": {
                "period_count": len(series),
                "total_sum": round(sum(item["sum"] for item in series), 6),
                "total_rows": sum(item["row_count"] for item in series),
            },
            "trend": summarize_trend(series),
            "series": series,
            "limitations": [
                "Analise temporal baseline; nao ajusta sazonalidade complexa.",
                "Periodos ausentes so aparecem quando existem linhas na fonte.",
            ],
        }

    def compare_periods(
        self,
        source: str,
        date_column: str,
        metric_column: str,
        baseline_start: str,
        baseline_end: str,
        comparison_start: str,
        comparison_end: str,
    ) -> dict[str, Any]:
        dataset = self.load_dataset(source)
        baseline = summarize_date_range(
            dataset,
            date_column,
            metric_column,
            parse_date_value(baseline_start),
            parse_date_value(baseline_end),
        )
        comparison = summarize_date_range(
            dataset,
            date_column,
            metric_column,
            parse_date_value(comparison_start),
            parse_date_value(comparison_end),
        )
        absolute = comparison["sum"] - baseline["sum"]
        percent = None if baseline["sum"] == 0 else round((absolute / baseline["sum"]) * 100, 6)
        return {
            "dataset": self.dataset_summary(dataset),
            "date_column": date_column,
            "metric_column": metric_column,
            "baseline": baseline,
            "comparison": comparison,
            "delta": {
                "absolute": round(absolute, 6),
                "percent": percent,
                "direction": "up" if absolute > 0 else "down" if absolute < 0 else "flat",
            },
        }

    def analyze_cohorts(
        self,
        source: str,
        cohort_column: str,
        event_date_column: str,
        metric_column: str | None = None,
    ) -> dict[str, Any]:
        dataset = self.load_dataset(source)
        for column in (cohort_column, event_date_column):
            if column not in dataset.columns:
                raise DataScientistError(f"column not found: {column}")
        if metric_column and metric_column not in dataset.columns:
            raise DataScientistError(f"metric column not found: {metric_column}")
        grouped: dict[str, dict[str, list[dict[str, str]]]] = defaultdict(lambda: defaultdict(list))
        for row in dataset.rows:
            cohort_date = parse_date_value(row.get(cohort_column, ""))
            event_date = parse_date_value(row.get(event_date_column, ""))
            if not cohort_date or not event_date:
                continue
            age = max(0, (event_date - cohort_date).days)
            grouped[cohort_date.isoformat()][str(age)].append(row)
        cohorts = []
        for cohort, periods in sorted(grouped.items()):
            cohort_periods = {}
            for period, rows in sorted(periods.items(), key=lambda item: int(item[0])):
                item: dict[str, Any] = {"row_count": len(rows)}
                if metric_column:
                    metric_values = [
                        float(value)
                        for value in (to_decimal(row.get(metric_column, "")) for row in rows)
                        if value is not None
                    ]
                    item["metric"] = summarize_metric(metric_values)
                cohort_periods[period] = item
            cohorts.append({"cohort": cohort, "periods": cohort_periods})
        return {
            "dataset": self.dataset_summary(dataset),
            "cohort_column": cohort_column,
            "event_date_column": event_date_column,
            "metric_column": metric_column,
            "summary": {
                "cohort_count": len(cohorts),
                "period_unit": "days_since_cohort",
            },
            "cohorts": cohorts,
        }

    def detect_anomalies(
        self,
        source: str,
        date_column: str,
        metric_column: str,
        granularity: str = "day",
        threshold: float = 2.0,
    ) -> dict[str, Any]:
        dataset = self.load_dataset(source)
        series = aggregate_time_series(dataset, date_column, metric_column, granularity)
        values = [item["sum"] for item in series]
        mean = statistics.fmean(values) if values else 0.0
        stdev = statistics.pstdev(values) if len(values) > 1 else 0.0
        anomalies = []
        if stdev > 0:
            for item in series:
                zscore = (item["sum"] - mean) / stdev
                if abs(zscore) > threshold:
                    anomalies.append(
                        {
                            "period": item["period"],
                            "value": item["sum"],
                            "zscore": round(zscore, 6),
                            "direction": "spike" if zscore > 0 else "drop",
                        }
                    )
        return {
            "dataset": self.dataset_summary(dataset),
            "date_column": date_column,
            "metric_column": metric_column,
            "granularity": granularity,
            "method": "zscore",
            "threshold": threshold,
            "baseline": {
                "mean": round(mean, 6),
                "stdev": round(stdev, 6),
            },
            "anomalies": anomalies,
            "series": series,
        }

    def forecast_series(
        self,
        source: str,
        date_column: str,
        metric_column: str,
        granularity: str = "day",
        periods: int = 3,
        window: int = 3,
    ) -> dict[str, Any]:
        dataset = self.load_dataset(source)
        series = aggregate_time_series(dataset, date_column, metric_column, granularity)
        if not series:
            raise DataScientistError("time series has no valid periods")
        periods = max(1, periods)
        window = max(1, window)
        historical = [item["sum"] for item in series]
        last_period = parse_period_label(series[-1]["period"], granularity)
        forecast = []
        generated_values: list[float] = []
        for index in range(1, periods + 1):
            basis = (historical + generated_values)[-window:]
            value = round(statistics.fmean(basis), 6)
            generated_values.append(value)
            period_date = advance_period(last_period, granularity, index)
            forecast.append({"period": format_period(period_date, granularity), "forecast": value})
        return {
            "dataset": self.dataset_summary(dataset),
            "date_column": date_column,
            "metric_column": metric_column,
            "granularity": granularity,
            "method": "moving_average",
            "window": window,
            "historical": series,
            "forecast": forecast,
            "limitations": [
                "Forecast baseline por media movel; nao modela sazonalidade ou eventos externos.",
            ],
        }

    def test_hypothesis(
        self,
        source: str,
        test_type: str,
        group_column: str,
        group_a: str,
        group_b: str,
        metric_column: str,
        alpha: float = 0.05,
    ) -> dict[str, Any]:
        if test_type != "mean-difference":
            raise DataScientistError(f"unsupported test type: {test_type}")
        dataset = self.load_dataset(source)
        left_values, right_values = grouped_metric_values(dataset, group_column, group_a, group_b, metric_column)
        left_summary = summarize_metric(left_values)
        right_summary = summarize_metric(right_values)
        standard_error = mean_difference_standard_error(left_values, right_values)
        difference = right_summary["mean"] - left_summary["mean"]
        zscore = None if standard_error == 0 else difference / standard_error
        p_value = None if zscore is None else round(two_tailed_normal_p_value(zscore), 8)
        return {
            "dataset": self.dataset_summary(dataset),
            "test_type": test_type,
            "alpha": alpha,
            "metric_column": metric_column,
            "group_column": group_column,
            "groups": {
                group_a: left_summary,
                group_b: right_summary,
            },
            "difference": round(difference, 6),
            "standard_error": round(standard_error, 6),
            "zscore": None if zscore is None else round(zscore, 6),
            "p_value": p_value,
            "significant": bool(p_value is not None and p_value < alpha),
            "interpretation": explain_significance(p_value, alpha, None)["executive_summary"],
            "limitations": [
                "Teste usa aproximacao normal e nao substitui validacao estatistica completa.",
                "Significancia estatistica nao implica causalidade.",
            ],
        }

    def calculate_confidence_intervals(
        self,
        source: str,
        metric_column: str,
        confidence: float = 0.95,
    ) -> dict[str, Any]:
        dataset = self.load_dataset(source)
        if metric_column not in dataset.columns:
            raise DataScientistError(f"metric column not found: {metric_column}")
        values = [
            float(value)
            for value in (to_decimal(row.get(metric_column, "")) for row in dataset.rows)
            if value is not None
        ]
        if len(values) < 2:
            raise DataScientistError("confidence interval requires at least two numeric values")
        mean = statistics.fmean(values)
        stdev = statistics.stdev(values)
        zvalue = z_for_confidence(confidence)
        margin = zvalue * (stdev / math.sqrt(len(values)))
        return {
            "dataset": self.dataset_summary(dataset),
            "metric_column": metric_column,
            "confidence": confidence,
            "sample_size": len(values),
            "mean": round(mean, 6),
            "standard_deviation": round(stdev, 6),
            "standard_error": round(stdev / math.sqrt(len(values)), 6),
            "z_value": zvalue,
            "interval": {
                "lower": round(mean - margin, 6),
                "upper": round(mean + margin, 6),
                "margin": round(margin, 6),
            },
            "limitations": [
                "Intervalo usa aproximacao normal e pressupoe amostra representativa.",
            ],
        }

    def calculate_sample_size(
        self,
        baseline_rate: float,
        minimum_detectable_effect: float,
        alpha: float = 0.05,
        power: float = 0.8,
    ) -> dict[str, Any]:
        if not 0 < baseline_rate < 1:
            raise DataScientistError("--baseline-rate must be between 0 and 1")
        if not 0 < minimum_detectable_effect < 1:
            raise DataScientistError("--minimum-detectable-effect must be between 0 and 1")
        z_alpha = z_for_two_tailed_alpha(alpha)
        z_power = z_for_power(power)
        variance = baseline_rate * (1 - baseline_rate)
        per_group = math.ceil(2 * ((z_alpha + z_power) ** 2) * variance / (minimum_detectable_effect**2))
        return {
            "method": "two_proportions_normal_approx",
            "baseline_rate": baseline_rate,
            "minimum_detectable_effect": minimum_detectable_effect,
            "alpha": alpha,
            "power": power,
            "z_alpha": z_alpha,
            "z_power": z_power,
            "sample_size_per_group": per_group,
            "total_sample_size": per_group * 2,
            "limitations": [
                "Estimativa baseline para duas proporcoes balanceadas.",
                "Amostras reais podem exigir ajuste por perdas, segmentacao ou desenho experimental.",
            ],
        }

    def measure_effect_size(
        self,
        source: str,
        group_column: str,
        group_a: str,
        group_b: str,
        metric_column: str,
    ) -> dict[str, Any]:
        dataset = self.load_dataset(source)
        left_values, right_values = grouped_metric_values(dataset, group_column, group_a, group_b, metric_column)
        effect = cohens_d(left_values, right_values)
        return {
            "dataset": self.dataset_summary(dataset),
            "method": "cohens_d",
            "metric_column": metric_column,
            "group_column": group_column,
            "groups": {
                group_a: summarize_metric(left_values),
                group_b: summarize_metric(right_values),
            },
            "effect_size": round(effect, 6),
            "magnitude": effect_magnitude(effect),
            "limitations": [
                "Cohen's d resume diferenca padronizada de medias; nao mede causalidade.",
            ],
        }

    def explain_statistical_result(
        self,
        p_value: float,
        alpha: float = 0.05,
        effect_size: float | None = None,
    ) -> dict[str, Any]:
        return explain_significance(p_value, alpha, effect_size)

    def reconcile_spreadsheets(
        self,
        left: str,
        right: str,
        key: str,
        compare_columns: list[str] | None = None,
        numeric_tolerance: Decimal = Decimal("0"),
    ) -> dict[str, Any]:
        left_dataset = self.load_dataset(left)
        right_dataset = self.load_dataset(right)
        keys = [item.strip() for item in key.split(",") if item.strip()]
        if not keys:
            raise DataScientistError("inform at least one --key column")
        missing = [
            f"left.{column}"
            for column in keys
            if column not in left_dataset.columns
        ] + [
            f"right.{column}"
            for column in keys
            if column not in right_dataset.columns
        ]
        if missing:
            raise DataScientistError(f"missing key columns: {', '.join(missing)}")
        if compare_columns is None:
            compare_columns = [
                column
                for column in left_dataset.columns
                if column in right_dataset.columns and column not in keys
            ]
        left_index = index_rows(left_dataset.rows, keys)
        right_index = index_rows(right_dataset.rows, keys)
        matched: list[dict[str, Any]] = []
        mismatched: list[dict[str, Any]] = []
        missing_right: list[dict[str, Any]] = []
        missing_left: list[dict[str, Any]] = []
        duplicate_keys = []
        all_keys = sorted(set(left_index) | set(right_index))

        for composite_key in all_keys:
            left_rows = left_index.get(composite_key, [])
            right_rows = right_index.get(composite_key, [])
            key_label = key_to_label(composite_key)
            if len(left_rows) > 1 or len(right_rows) > 1:
                duplicate_keys.append(
                    {
                        "key": key_label,
                        "left_count": len(left_rows),
                        "right_count": len(right_rows),
                    }
                )
            if not left_rows:
                missing_left.append({"key": key_label, "right": mask_row(right_rows[0])})
                continue
            if not right_rows:
                missing_right.append({"key": key_label, "left": mask_row(left_rows[0])})
                continue
            differences = compare_rows(
                left_rows[0],
                right_rows[0],
                compare_columns,
                numeric_tolerance,
            )
            if differences:
                mismatched.append({"key": key_label, "differences": differences})
            else:
                matched.append({"key": key_label})

        return {
            "sources": {
                "left": self.dataset_summary(left_dataset),
                "right": self.dataset_summary(right_dataset),
            },
            "rules": {
                "key": keys,
                "compare_columns": compare_columns,
                "numeric_tolerance": str(numeric_tolerance),
            },
            "summary": {
                "matched_count": len(matched),
                "mismatched_count": len(mismatched),
                "missing_right_count": len(missing_right),
                "missing_left_count": len(missing_left),
                "duplicate_key_count": len(duplicate_keys),
            },
            "matched": matched,
            "mismatched": mismatched,
            "missing_right": missing_right,
            "missing_left": missing_left,
            "duplicate_keys": duplicate_keys,
        }

    def generate_reconciliation_report(
        self,
        left: str,
        right: str,
        key: str,
        compare_columns: list[str] | None,
        numeric_tolerance: Decimal,
        output: str | None = None,
    ) -> dict[str, Any]:
        payload = self.reconcile_spreadsheets(
            left=left,
            right=right,
            key=key,
            compare_columns=compare_columns,
            numeric_tolerance=numeric_tolerance,
        )
        content = render_reconciliation_report(payload)
        written = write_artifact(content, output)
        payload["report"] = {"format": "markdown", "output": written, "content": content}
        return payload

    def generate_data_report(self, source: str, output: str | None = None) -> dict[str, Any]:
        payload = self.profile_dataset(source)
        content = render_data_report(payload)
        written = write_artifact(content, output)
        return {"profile": payload, "report": {"format": "markdown", "output": written, "content": content}}

    def load_dataset(self, source: str) -> Dataset:
        path = Path(source).expanduser().resolve()
        if not path.exists():
            raise DataScientistError(f"source not found: {source}")
        if path.is_dir():
            rows: list[dict[str, str]] = []
            columns: list[str] = []
            files = [
                item
                for item in sorted(path.iterdir())
                if item.is_file() and item.suffix.lower() in {".csv", ".json", ".jsonl"}
            ]
            if not files:
                raise DataScientistError(f"directory has no supported dataset files: {source}")
            for item in files:
                dataset = self.load_dataset(str(item))
                columns = merge_columns(columns, dataset.columns)
                rows.extend(dataset.rows)
            return Dataset(
                source=str(path),
                name=path.name,
                format="directory",
                rows=rows,
                columns=columns,
                sha256=hash_directory(files),
            )
        suffix = path.suffix.lower()
        if suffix == ".csv":
            rows, columns = read_csv(path)
        elif suffix == ".json":
            rows, columns = read_json(path)
        elif suffix == ".jsonl":
            rows, columns = read_jsonl(path)
        elif suffix == ".xlsx":
            rows, columns = read_xlsx(path)
        else:
            raise DataScientistError(f"unsupported dataset format: {suffix or path.name}")
        return Dataset(
            source=str(path),
            name=path.name,
            format=suffix.lstrip("."),
            rows=rows,
            columns=columns,
            sha256=hash_file(path),
        )

    def dataset_summary(self, dataset: Dataset) -> dict[str, Any]:
        return {
            "source": dataset.source,
            "name": dataset.name,
            "format": dataset.format,
            "row_count": len(dataset.rows),
            "column_count": len(dataset.columns),
            "sha256": dataset.sha256,
        }


def read_csv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    with path.open(newline="", encoding="utf-8-sig") as file:
        sample = file.read(4096)
        file.seek(0)
        dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
        reader = csv.DictReader(file, dialect=dialect)
        columns = [column.strip() for column in (reader.fieldnames or [])]
        rows = [normalize_row(row, columns) for row in reader]
    return rows, columns


def read_json(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(raw, dict):
        for value in raw.values():
            if isinstance(value, list):
                raw = value
                break
    if not isinstance(raw, list):
        raise DataScientistError("json source must be an object with a list or a list of records")
    rows = [flatten_record(item) for item in raw if isinstance(item, dict)]
    columns = merge_columns([], *[list(row.keys()) for row in rows])
    return rows, columns


def read_jsonl(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    rows = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.strip():
            item = json.loads(line)
            if isinstance(item, dict):
                rows.append(flatten_record(item))
    columns = merge_columns([], *[list(row.keys()) for row in rows])
    return rows, columns


def read_xlsx(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise DataScientistError("xlsx support requires local dependency openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = next(iterator)
    except StopIteration:
        return [], []
    columns = [str(value).strip() if value is not None else f"column_{index + 1}" for index, value in enumerate(headers)]
    rows = []
    for values in iterator:
        row = {
            columns[index]: value_to_string(values[index] if index < len(values) else "")
            for index in range(len(columns))
        }
        rows.append(row)
    return rows, columns


def normalize_row(row: dict[str, Any], columns: list[str]) -> dict[str, str]:
    return {column: value_to_string(row.get(column, "")) for column in columns}


def flatten_record(record: dict[str, Any], prefix: str = "") -> dict[str, str]:
    flattened = {}
    for key, value in record.items():
        name = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            flattened.update(flatten_record(value, name))
        else:
            flattened[name] = value_to_string(value)
    return flattened


def value_to_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def profile_column(column: str, values: list[str]) -> dict[str, Any]:
    present = [value for value in values if value != ""]
    inferred = infer_type(present)
    frequency = Counter(present)
    profile: dict[str, Any] = {
        "inferred_type": inferred,
        "missing_count": len(values) - len(present),
        "unique_count": len(frequency),
        "sample_values": [mask_value(column, value) for value in list(dict.fromkeys(present))[:5]],
        "top_values": [
            {"value": mask_value(column, value), "count": count}
            for value, count in frequency.most_common(5)
        ],
    }
    if inferred == "number" and present:
        numbers = [float(to_decimal(value)) for value in present if to_decimal(value) is not None]
        if numbers:
            profile["numeric"] = {
                "min": min(numbers),
                "max": max(numbers),
                "mean": round(statistics.fmean(numbers), 4),
                "median": round(statistics.median(numbers), 4),
            }
    return profile


def resolve_numeric_columns(dataset: Dataset, columns: list[str] | None) -> list[str]:
    selected = columns or dataset.columns
    missing = [column for column in selected if column not in dataset.columns]
    if missing:
        raise DataScientistError(f"columns not found: {', '.join(missing)}")
    numeric = []
    for column in selected:
        values = [row.get(column, "") for row in dataset.rows if row.get(column, "") != ""]
        if infer_type(values) == "number":
            numeric.append(column)
    return numeric


def numeric_values_with_rows(rows: list[dict[str, str]], column: str) -> list[dict[str, Any]]:
    values = []
    for index, row in enumerate(rows, start=1):
        value = to_decimal(row.get(column, ""))
        if value is not None:
            values.append({"row_number": index, "value": float(value), "raw": row.get(column, "")})
    return values


def detect_column_outliers(values: list[dict[str, Any]], method: str, threshold: float) -> dict[str, Any]:
    method = method.lower()
    numbers = [item["value"] for item in values]
    outlier_rows: dict[int, dict[str, Any]] = {}
    limits: dict[str, Any] = {}
    if method in {"iqr", "both"}:
        q1 = percentile(numbers, 25)
        q3 = percentile(numbers, 75)
        iqr = q3 - q1
        lower = q1 - (1.5 * iqr)
        upper = q3 + (1.5 * iqr)
        limits["iqr"] = {
            "q1": round(q1, 6),
            "q3": round(q3, 6),
            "lower": round(lower, 6),
            "upper": round(upper, 6),
        }
        for item in values:
            if item["value"] < lower or item["value"] > upper:
                outlier_rows[item["row_number"]] = {
                    "row_number": item["row_number"],
                    "value": item["value"],
                    "methods": ["iqr"],
                }
    if method in {"zscore", "both"}:
        mean = statistics.fmean(numbers)
        stdev = statistics.pstdev(numbers)
        limits["zscore"] = {
            "mean": round(mean, 6),
            "stdev": round(stdev, 6),
            "threshold": threshold,
        }
        if stdev > 0:
            for item in values:
                zscore = abs((item["value"] - mean) / stdev)
                if zscore > threshold:
                    existing = outlier_rows.setdefault(
                        item["row_number"],
                        {"row_number": item["row_number"], "value": item["value"], "methods": []},
                    )
                    existing["methods"].append("zscore")
                    existing["zscore"] = round(zscore, 6)
    outliers = sorted(outlier_rows.values(), key=lambda item: item["row_number"])
    return {
        "method": method,
        "outlier_count": len(outliers),
        "outlier_ratio": round(len(outliers) / len(values), 6) if values else 0.0,
        "limits": limits,
        "outliers": outliers[:50],
    }


def percentile(values: list[float], percent: float) -> float:
    ordered = sorted(values)
    if not ordered:
        return 0.0
    if len(ordered) == 1:
        return ordered[0]
    rank = (len(ordered) - 1) * (percent / 100)
    lower = math.floor(rank)
    upper = math.ceil(rank)
    if lower == upper:
        return ordered[int(rank)]
    weight = rank - lower
    return ordered[lower] * (1 - weight) + ordered[upper] * weight


def paired_numeric_values(rows: list[dict[str, str]], left: str, right: str) -> list[tuple[float, float]]:
    pairs = []
    for row in rows:
        left_value = to_decimal(row.get(left, ""))
        right_value = to_decimal(row.get(right, ""))
        if left_value is not None and right_value is not None:
            pairs.append((float(left_value), float(right_value)))
    return pairs


def pearson_correlation(left_values: list[float], right_values: list[float]) -> float | None:
    if len(left_values) != len(right_values) or len(left_values) < 2:
        return None
    left_mean = statistics.fmean(left_values)
    right_mean = statistics.fmean(right_values)
    numerator = sum((left - left_mean) * (right - right_mean) for left, right in zip(left_values, right_values))
    left_denominator = math.sqrt(sum((left - left_mean) ** 2 for left in left_values))
    right_denominator = math.sqrt(sum((right - right_mean) ** 2 for right in right_values))
    denominator = left_denominator * right_denominator
    if denominator == 0:
        return None
    return numerator / denominator


def correlation_strength(value: float) -> str:
    absolute = abs(value)
    if absolute >= 0.8:
        return "strong"
    if absolute >= 0.5:
        return "moderate"
    if absolute >= 0.3:
        return "weak"
    return "very_weak"


def summarize_metric(values: list[float]) -> dict[str, Any]:
    if not values:
        return {"count": 0, "sum": 0.0, "mean": None, "min": None, "max": None}
    return {
        "count": len(values),
        "sum": round(sum(values), 6),
        "mean": round(statistics.fmean(values), 6),
        "min": min(values),
        "max": max(values),
    }


def aggregate_time_series(
    dataset: Dataset,
    date_column: str,
    metric_column: str,
    granularity: str,
) -> list[dict[str, Any]]:
    for column in (date_column, metric_column):
        if column not in dataset.columns:
            raise DataScientistError(f"column not found: {column}")
    grouped: dict[str, list[float]] = defaultdict(list)
    row_counts: dict[str, int] = defaultdict(int)
    for row in dataset.rows:
        parsed_date = parse_date_value(row.get(date_column, ""))
        metric_value = to_decimal(row.get(metric_column, ""))
        if not parsed_date or metric_value is None:
            continue
        period = format_period(bucket_date(parsed_date, granularity), granularity)
        grouped[period].append(float(metric_value))
        row_counts[period] += 1
    series = []
    for period in sorted(grouped, key=lambda item: parse_period_label(item, granularity)):
        values = grouped[period]
        series.append(
            {
                "period": period,
                "row_count": row_counts[period],
                "sum": round(sum(values), 6),
                "mean": round(statistics.fmean(values), 6),
                "min": min(values),
                "max": max(values),
            }
        )
    return series


def summarize_trend(series: list[dict[str, Any]]) -> dict[str, Any]:
    if len(series) < 2:
        return {"direction": "insufficient_data", "delta": None, "percent": None}
    first = series[0]["sum"]
    last = series[-1]["sum"]
    delta = last - first
    percent = None if first == 0 else round((delta / first) * 100, 6)
    return {
        "direction": "up" if delta > 0 else "down" if delta < 0 else "flat",
        "delta": round(delta, 6),
        "percent": percent,
        "first_period": series[0]["period"],
        "last_period": series[-1]["period"],
    }


def summarize_date_range(
    dataset: Dataset,
    date_column: str,
    metric_column: str,
    start: date | None,
    end: date | None,
) -> dict[str, Any]:
    if not start or not end:
        raise DataScientistError("invalid date range")
    if start > end:
        raise DataScientistError("range start must be before range end")
    values = []
    row_count = 0
    for row in dataset.rows:
        parsed_date = parse_date_value(row.get(date_column, ""))
        metric_value = to_decimal(row.get(metric_column, ""))
        if parsed_date and metric_value is not None and start <= parsed_date <= end:
            values.append(float(metric_value))
            row_count += 1
    metric = summarize_metric(values)
    return {
        "start": start.isoformat(),
        "end": end.isoformat(),
        "row_count": row_count,
        **metric,
    }


def parse_date_value(value: str) -> date | None:
    text = value.strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            continue
    return None


def bucket_date(value: date, granularity: str) -> date:
    normalized = granularity.lower()
    if normalized == "day":
        return value
    if normalized == "week":
        return value - timedelta(days=value.weekday())
    if normalized == "month":
        return value.replace(day=1)
    raise DataScientistError(f"unsupported granularity: {granularity}")


def format_period(value: date, granularity: str) -> str:
    normalized = granularity.lower()
    if normalized == "month":
        return value.strftime("%Y-%m")
    return value.isoformat()


def parse_period_label(value: str, granularity: str) -> date:
    if granularity.lower() == "month":
        return datetime.strptime(value, "%Y-%m").date()
    parsed = parse_date_value(value)
    if not parsed:
        raise DataScientistError(f"invalid period: {value}")
    return parsed


def advance_period(value: date, granularity: str, amount: int) -> date:
    normalized = granularity.lower()
    if normalized == "day":
        return value + timedelta(days=amount)
    if normalized == "week":
        return value + timedelta(weeks=amount)
    if normalized == "month":
        month = value.month - 1 + amount
        year = value.year + month // 12
        month = month % 12 + 1
        return date(year, month, 1)
    raise DataScientistError(f"unsupported granularity: {granularity}")


def grouped_metric_values(
    dataset: Dataset,
    group_column: str,
    group_a: str,
    group_b: str,
    metric_column: str,
) -> tuple[list[float], list[float]]:
    for column in (group_column, metric_column):
        if column not in dataset.columns:
            raise DataScientistError(f"column not found: {column}")
    left_values = []
    right_values = []
    for row in dataset.rows:
        metric = to_decimal(row.get(metric_column, ""))
        if metric is None:
            continue
        group_value = row.get(group_column, "")
        if group_value == group_a:
            left_values.append(float(metric))
        elif group_value == group_b:
            right_values.append(float(metric))
    if len(left_values) < 2 or len(right_values) < 2:
        raise DataScientistError("group comparison requires at least two numeric values per group")
    return left_values, right_values


def mean_difference_standard_error(left_values: list[float], right_values: list[float]) -> float:
    left_variance = statistics.variance(left_values)
    right_variance = statistics.variance(right_values)
    return math.sqrt((left_variance / len(left_values)) + (right_variance / len(right_values)))


def two_tailed_normal_p_value(zscore: float) -> float:
    return max(0.0, min(1.0, 2 * (1 - normal_cdf(abs(zscore)))))


def normal_cdf(value: float) -> float:
    return 0.5 * (1 + math.erf(value / math.sqrt(2)))


def z_for_confidence(confidence: float) -> float:
    lookup = {
        0.8: 1.281552,
        0.9: 1.644854,
        0.95: 1.959964,
        0.98: 2.326348,
        0.99: 2.575829,
    }
    rounded = round(confidence, 2)
    if rounded in lookup:
        return lookup[rounded]
    raise DataScientistError(f"unsupported confidence level: {confidence}")


def z_for_two_tailed_alpha(alpha: float) -> float:
    lookup = {
        0.1: 1.644854,
        0.05: 1.959964,
        0.01: 2.575829,
    }
    rounded = round(alpha, 2)
    if rounded in lookup:
        return lookup[rounded]
    raise DataScientistError(f"unsupported alpha: {alpha}")


def z_for_power(power: float) -> float:
    lookup = {
        0.7: 0.524401,
        0.8: 0.841621,
        0.85: 1.036433,
        0.9: 1.281552,
        0.95: 1.644854,
    }
    rounded = round(power, 2)
    if rounded in lookup:
        return lookup[rounded]
    raise DataScientistError(f"unsupported power: {power}")


def cohens_d(left_values: list[float], right_values: list[float]) -> float:
    left_variance = statistics.variance(left_values)
    right_variance = statistics.variance(right_values)
    pooled_variance = (
        ((len(left_values) - 1) * left_variance + (len(right_values) - 1) * right_variance)
        / (len(left_values) + len(right_values) - 2)
    )
    pooled_stdev = math.sqrt(pooled_variance)
    if pooled_stdev == 0:
        return 0.0
    return (statistics.fmean(right_values) - statistics.fmean(left_values)) / pooled_stdev


def effect_magnitude(effect_size: float | None) -> str | None:
    if effect_size is None:
        return None
    absolute = abs(effect_size)
    if absolute >= 0.8:
        return "large"
    if absolute >= 0.5:
        return "medium"
    if absolute >= 0.2:
        return "small"
    return "negligible"


def explain_significance(p_value: float | None, alpha: float, effect_size: float | None) -> dict[str, Any]:
    significant = bool(p_value is not None and p_value < alpha)
    magnitude = effect_magnitude(effect_size)
    if significant:
        summary = "Resultado estatisticamente significativo para o alpha informado."
    else:
        summary = "Resultado nao e estatisticamente significativo para o alpha informado."
    if magnitude:
        summary += f" O tamanho de efeito e {magnitude}."
    return {
        "p_value": p_value,
        "alpha": alpha,
        "effect_size": effect_size,
        "effect_magnitude": magnitude,
        "significant": significant,
        "executive_summary": summary,
        "technical_notes": [
            "p-valor menor que alpha indica evidencia contra a hipotese nula, nao prova causalidade.",
            "Tamanho de efeito ajuda a avaliar relevancia pratica alem da significancia.",
        ],
    }


def build_hypotheses(
    profile: dict[str, Any],
    outliers: dict[str, Any],
    correlations: dict[str, Any],
    segments: dict[str, Any],
    target_column: str | None,
) -> list[str]:
    hypotheses = []
    if profile["quality"]["columns_with_missing_values"]:
        hypotheses.append("Investigar impacto de valores ausentes nas colunas criticas.")
    outlier_columns = [
        column
        for column, details in outliers.get("columns", {}).items()
        if details.get("outlier_count", 0) > 0
    ]
    if outlier_columns:
        hypotheses.append(f"Validar se outliers em {', '.join(outlier_columns[:5])} representam erro, fraude ou evento raro.")
    if correlations.get("top_correlations"):
        pair = correlations["top_correlations"][0]
        hypotheses.append(
            f"Explorar a relacao entre {pair['left']} e {pair['right']} antes de usar como explicacao causal."
        )
    if target_column and correlations.get("target_correlations"):
        top = correlations["target_correlations"][0]
        hypotheses.append(f"Avaliar se {top['column']} ajuda a explicar variacao em {target_column}.")
    if segments.get("segments"):
        hypotheses.append("Comparar segmentos dominantes para identificar diferencas de comportamento ou qualidade.")
    return hypotheses or ["Coletar mais contexto de negocio para formular hipoteses acionaveis."]


def infer_type(values: list[str]) -> str:
    if not values:
        return "empty"
    number_count = sum(1 for value in values if to_decimal(value) is not None)
    date_count = sum(1 for value in values if looks_like_date(value))
    bool_count = sum(1 for value in values if value.lower() in {"true", "false", "sim", "nao", "não", "0", "1"})
    threshold = max(1, math.ceil(len(values) * 0.8))
    if number_count >= threshold:
        return "number"
    if date_count >= threshold:
        return "date"
    if bool_count >= threshold:
        return "boolean"
    return "text"


def looks_like_date(value: str) -> bool:
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            datetime.strptime(value[:19], fmt)
            return True
        except ValueError:
            continue
    return False


def to_decimal(value: str) -> Decimal | None:
    normalized = value.strip().replace("R$", "").replace(" ", "")
    if normalized.count(",") == 1 and normalized.count(".") > 1:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif normalized.count(",") == 1 and normalized.count(".") == 0:
        normalized = normalized.replace(",", ".")
    try:
        return Decimal(normalized)
    except (InvalidOperation, ValueError):
        return None


def detect_sensitive(dataset: Dataset) -> dict[str, Any]:
    columns: dict[str, list[str]] = {}
    examples: dict[str, list[str]] = {}
    for column in dataset.columns:
        kinds = set()
        lowered = normalize_name(column)
        for pattern, kind in SENSITIVE_NAME_PATTERNS.items():
            if pattern in lowered:
                kinds.add(kind)
        samples = [row.get(column, "") for row in dataset.rows[:100] if row.get(column, "")]
        if "cpf" in lowered or any(is_formatted_cpf(value) for value in samples):
            kinds.add("cpf")
        if any(CNPJ_RE.search(value) for value in samples):
            kinds.add("cnpj")
        if any(EMAIL_RE.search(value) for value in samples):
            kinds.add("email")
        if lowered not in {"amount", "value", "valor"} and any(PHONE_RE.fullmatch(value) for value in samples):
            kinds.add("phone")
        if kinds:
            columns[column] = sorted(kinds)
            examples[column] = [mask_value(column, value) for value in samples[:3]]
    return {
        "has_sensitive_data": bool(columns),
        "columns": columns,
        "masked_examples": examples,
    }


def infer_probable_keys(dataset: Dataset, column_profiles: dict[str, dict[str, Any]]) -> list[str]:
    row_count = len(dataset.rows)
    keys = []
    for column, profile in column_profiles.items():
        lowered = normalize_name(column)
        if row_count and profile["unique_count"] == row_count and profile["missing_count"] == 0:
            keys.append(column)
        elif lowered in {"id", "codigo", "code", "cpf", "cnpj"} and profile["missing_count"] == 0:
            keys.append(column)
    return keys[:10]


def count_duplicate_rows(rows: list[dict[str, str]], columns: list[str]) -> int:
    seen: set[tuple[str, ...]] = set()
    duplicates = 0
    for row in rows:
        signature = tuple(row.get(column, "") for column in columns)
        if signature in seen:
            duplicates += 1
        else:
            seen.add(signature)
    return duplicates


def count_empty_rows(rows: list[dict[str, str]]) -> int:
    return sum(1 for row in rows if all(value == "" for value in row.values()))


def quality_score(row_count: int, column_profiles: dict[str, dict[str, Any]], duplicate_row_count: int) -> int:
    if row_count == 0:
        return 0
    total_cells = row_count * max(len(column_profiles), 1)
    missing = sum(profile["missing_count"] for profile in column_profiles.values())
    duplicate_penalty = duplicate_row_count / row_count
    missing_penalty = missing / total_cells if total_cells else 0
    return max(0, min(100, round(100 - (missing_penalty * 50) - (duplicate_penalty * 30))))


def index_rows(rows: list[dict[str, str]], keys: list[str]) -> dict[tuple[str, ...], list[dict[str, str]]]:
    indexed: dict[tuple[str, ...], list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        indexed[tuple(normalize_key_value(row.get(column, "")) for column in keys)].append(row)
    return indexed


def compare_rows(
    left: dict[str, str],
    right: dict[str, str],
    columns: list[str],
    numeric_tolerance: Decimal,
) -> list[dict[str, Any]]:
    differences = []
    for column in columns:
        left_value = left.get(column, "")
        right_value = right.get(column, "")
        left_number = to_decimal(left_value)
        right_number = to_decimal(right_value)
        if left_number is not None and right_number is not None:
            delta = abs(left_number - right_number)
            if delta <= numeric_tolerance:
                continue
            reason = "numeric_difference"
        elif normalize_key_value(left_value) == normalize_key_value(right_value):
            continue
        else:
            delta = None
            reason = "value_difference"
        differences.append(
            {
                "column": column,
                "left": mask_value(column, left_value),
                "right": mask_value(column, right_value),
                "reason": reason,
                "delta": str(delta) if delta is not None else None,
            }
        )
    return differences


def key_to_label(key: tuple[str, ...]) -> str:
    return "|".join(key)


def normalize_key_value(value: str) -> str:
    stripped = value.strip()
    digits = re.sub(r"\D", "", stripped)
    if len(digits) in {11, 14}:
        return digits
    return stripped.casefold()


def normalize_name(value: str) -> str:
    return value.strip().casefold().replace("_", " ").replace("-", " ")


def is_formatted_cpf(value: str) -> bool:
    return bool(re.fullmatch(r"\d{3}\.?\d{3}\.?\d{3}-\d{2}", value.strip()))


def mask_row(row: dict[str, str]) -> dict[str, str]:
    return {column: mask_value(column, value) for column, value in row.items()}


def mask_value(column: str, value: str) -> str:
    text = value_to_string(value)
    lowered = normalize_name(column)
    if not text:
        return text
    digits = re.sub(r"\D", "", text)
    if "cpf" in lowered or len(digits) == 11:
        return f"***{digits[-2:]}" if digits else "***"
    if "cnpj" in lowered or len(digits) == 14:
        return f"***{digits[-4:]}" if digits else "***"
    if "email" in lowered or EMAIL_RE.fullmatch(text):
        local, _, domain = text.partition("@")
        return f"{local[:1]}***@{domain}" if domain else "***"
    if "telefone" in lowered or "phone" in lowered or "celular" in lowered:
        return f"***{digits[-4:]}" if digits else "***"
    return text


def merge_columns(base: list[str], *column_sets: list[str]) -> list[str]:
    columns = list(base)
    for column_set in column_sets:
        for column in column_set:
            if column not in columns:
                columns.append(column)
    return columns


def hash_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def hash_directory(files: list[Path]) -> str:
    digest = hashlib.sha256()
    for path in files:
        digest.update(path.name.encode("utf-8"))
        digest.update(hash_file(path).encode("utf-8"))
    return digest.hexdigest()


def write_artifact(content: str, output: str | None) -> str | None:
    if not output:
        return None
    path = Path(output).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")
    return str(path)


def render_data_report(payload: dict[str, Any]) -> str:
    dataset = payload["dataset"]
    quality = payload["quality"]
    sensitive = payload["sensitive_data"]
    lines = [
        "# Relatorio de Dados",
        "",
        "## Fonte",
        "",
        f"- Arquivo: {dataset['name']}",
        f"- Formato: {dataset['format']}",
        f"- Linhas: {dataset['row_count']}",
        f"- Colunas: {dataset['column_count']}",
        f"- SHA-256: {dataset['sha256']}",
        "",
        "## Qualidade",
        "",
        f"- Score: {quality['quality_score']}",
        f"- Linhas duplicadas: {quality['duplicate_row_count']}",
        f"- Linhas vazias: {quality['empty_row_count']}",
        f"- Colunas com nulos: {', '.join(quality['columns_with_missing_values']) or '-'}",
        "",
        "## Dados sensiveis",
        "",
        f"- Detectados: {'sim' if sensitive['has_sensitive_data'] else 'nao'}",
        f"- Colunas: {', '.join(sensitive['columns'].keys()) or '-'}",
        "",
        "## Colunas",
        "",
        "| Coluna | Tipo | Nulos | Unicos |",
        "|---|---|---:|---:|",
    ]
    for column, details in payload["columns"].items():
        lines.append(
            f"| {column} | {details['inferred_type']} | {details['missing_count']} | {details['unique_count']} |"
        )
    return "\n".join(lines).rstrip() + "\n"


def render_reconciliation_report(payload: dict[str, Any]) -> str:
    summary = payload["summary"]
    rules = payload["rules"]
    lines = [
        "# Relatorio de Conciliacao",
        "",
        "## Regras",
        "",
        f"- Chaves: {', '.join(rules['key'])}",
        f"- Colunas comparadas: {', '.join(rules['compare_columns']) or '-'}",
        f"- Tolerancia numerica: {rules['numeric_tolerance']}",
        "",
        "## Resumo",
        "",
        f"- Registros conciliados: {summary['matched_count']}",
        f"- Registros divergentes: {summary['mismatched_count']}",
        f"- Ausentes na direita: {summary['missing_right_count']}",
        f"- Ausentes na esquerda: {summary['missing_left_count']}",
        f"- Chaves duplicadas: {summary['duplicate_key_count']}",
        "",
        "## Divergencias",
        "",
    ]
    if not payload["mismatched"]:
        lines.append("- Nenhuma divergencia de valores encontrada.")
    for item in payload["mismatched"][:50]:
        fields = ", ".join(diff["column"] for diff in item["differences"])
        lines.append(f"- Chave {item['key']}: {fields}")
    return "\n".join(lines).rstrip() + "\n"
