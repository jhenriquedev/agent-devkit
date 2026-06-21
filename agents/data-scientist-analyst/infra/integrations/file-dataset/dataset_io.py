#!/usr/bin/env python3
"""Dataset IO helpers for local tabular files."""

from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path
from typing import Any

from dataset_models import DataScientistError


def read_csv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    with path.open(newline="", encoding="utf-8-sig") as file:
        sample = file.read(4096)
        file.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(file, dialect=dialect)
        columns = [column.strip() for column in (reader.fieldnames or [])]
        rows = [normalize_row(row, columns) for row in reader]
    return rows, columns


def read_json(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(raw, dict):
        for value in raw.values():
            if isinstance(value, list):
                raw = value
                break
    if not isinstance(raw, list):
        raise DataScientistError("json source must be an object with a list or a list of records")
    rows = [flatten_record(item) for item in raw if isinstance(item, dict)]
    columns = merge_columns([], *[list(row.keys()) for row in rows])
    return rows, columns


def read_jsonl(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    rows = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.strip():
            item = json.loads(line)
            if isinstance(item, dict):
                rows.append(flatten_record(item))
    columns = merge_columns([], *[list(row.keys()) for row in rows])
    return rows, columns


def read_xlsx(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise DataScientistError("xlsx support requires local dependency openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = next(iterator)
    except StopIteration:
        return [], []
    columns = [str(value).strip() if value is not None else f"column_{index + 1}" for index, value in enumerate(headers)]
    rows = []
    for values in iterator:
        row = {
            columns[index]: value_to_string(values[index] if index < len(values) else "")
            for index in range(len(columns))
        }
        rows.append(row)
    return rows, columns


def normalize_row(row: dict[str, Any], columns: list[str]) -> dict[str, str]:
    return {column: value_to_string(row.get(column, "")) for column in columns}


def flatten_record(record: dict[str, Any], prefix: str = "") -> dict[str, str]:
    flattened = {}
    for key, value in record.items():
        name = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            flattened.update(flatten_record(value, name))
        elif isinstance(value, list):
            flattened[name] = json.dumps(value, ensure_ascii=False)
        else:
            flattened[name] = value_to_string(value)
    return flattened


def value_to_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def merge_columns(base: list[str], *column_sets: list[str]) -> list[str]:
    columns = list(base)
    for column_set in column_sets:
        for column in column_set:
            if column not in columns:
                columns.append(column)
    return columns


def hash_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def hash_directory(files: list[Path]) -> str:
    digest = hashlib.sha256()
    for path in files:
        digest.update(path.name.encode("utf-8"))
        digest.update(hash_file(path).encode("utf-8"))
    return digest.hexdigest()
